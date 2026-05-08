#!/usr/bin/env python3
"""Convert UNCTAD ISDS Navigator Excel → isds.db (SQLite)."""
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///

import sqlite3
import re
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "UNCTAD-ISDS-Navigator-data-set-31December2023.xlsx"
DB = Path(__file__).parent / "isds.db"

HEADER_ROW = 12  # 1-indexed
DATA_START = 13

COL_MAP = {
    "NO.":                                                          "no",
    "YEAR OF INITIATION":                                           "year",
    "SHORT CASE NAME":                                              "short_name",
    "FULL CASE NAME":                                               "full_name",
    "APPLICABLE IIA":                                               "applicable_iia",
    "ARBITRAL RULES":                                               "arbitral_rules",
    "ADMINISTERING INSTITUTION":                                    "institution",
    "STATUS/OUTCOME OF ORIGINAL PROCEEDINGS":                       "status",
    "RESPONDENT STATE":                                             "respondent_state",
    "HOME STATE OF INVESTOR":                                       "home_state",
    "ECONOMIC SECTOR":                                              "sector",
    "ECONOMIC SUBSECTOR":                                           "subsector",
    "SUMMARY OF THE DISPUTE":                                       "summary",
    " DETAILS OF INVESTMENT":                                       "investment_details",
    "ARBITRATORS":                                                  "arbitrators",
    "DECISIONS":                                                    "decisions",
    "INDIVIDUAL OPINIONS DETAILS":                                  "individual_opinions",
    "AMOUNT CLAIMED (EXPRESSED IN MILLIONS)":                       "amount_claimed_m",
    "AMOUNT AWARDED (OR SETTLED FOR) (EXPRESSED IN MILLIONS)":      "amount_awarded_m",
    "IIA BREACHES ALLEGED":                                         "breaches_alleged",
    "IIA BREACHES FOUND":                                           "breaches_found",
    "FOLLOW-ON PROCEEDING TYPE":                                    "followon_type",
    "FOLLOW-ON PROCEEDING STATUS":                                  "followon_status",
    "FOLLOW-ON DECISIONS":                                          "followon_decisions",
    "FOLLOW-ON INDIVIDUAL OPINIONS":                                "followon_opinions",
    "ICSID ANNULMENT COMMITTEE MEMBERS":                            "annulment_committee",
    "LINK TO ITALAW'S CASE PAGE":                                   "italaw_link",
    "LINKS TO BACKGROUND SOURCES":                                  "background_sources",
}

CREATE_SQL = """
CREATE TABLE IF NOT EXISTS cases (
    no                  INTEGER PRIMARY KEY,
    year                INTEGER,
    short_name          TEXT,
    full_name           TEXT,
    applicable_iia      TEXT,
    arbitral_rules      TEXT,
    institution         TEXT,
    status              TEXT,
    respondent_state    TEXT,
    home_state          TEXT,
    sector              TEXT,
    subsector           TEXT,
    summary             TEXT,
    investment_details  TEXT,
    arbitrators         TEXT,
    decisions           TEXT,
    individual_opinions TEXT,
    amount_claimed_m    TEXT,
    amount_awarded_m    TEXT,
    breaches_alleged    TEXT,
    breaches_found      TEXT,
    followon_type       TEXT,
    followon_status     TEXT,
    followon_decisions  TEXT,
    followon_opinions   TEXT,
    annulment_committee TEXT,
    italaw_link         TEXT,
    background_sources  TEXT
)
"""

CREATE_FTS = """
CREATE VIRTUAL TABLE IF NOT EXISTS cases_fts USING fts5(
    short_name, full_name, applicable_iia, respondent_state,
    home_state, sector, status, summary,
    content='cases', content_rowid='no'
)
"""

POPULATE_FTS = """
INSERT INTO cases_fts(rowid, short_name, full_name, applicable_iia,
    respondent_state, home_state, sector, status, summary)
SELECT no, short_name, full_name, applicable_iia,
    respondent_state, home_state, sector, status, summary
FROM cases
"""


def load():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb["Source - UNCTAD ISDS Navigator"]

    # Read headers from row 12
    raw_headers = [cell.value for cell in next(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW)
    )]

    col_indices = {}
    for i, h in enumerate(raw_headers):
        if h in COL_MAP:
            col_indices[COL_MAP[h]] = i

    rows = []
    for raw in ws.iter_rows(min_row=DATA_START, values_only=True):
        if all(v is None for v in raw):
            continue
        row = {name: raw[idx] for name, idx in col_indices.items()}
        # Normalise cell values
        for k, v in row.items():
            if v is None:
                row[k] = None
            else:
                row[k] = str(v).strip() if not isinstance(v, (int, float)) else v
        rows.append(row)

    wb.close()
    return rows


def build_db(rows):
    DB.unlink(missing_ok=True)
    con = sqlite3.connect(DB)
    con.execute(CREATE_SQL)
    cols = list(COL_MAP.values())
    placeholders = ", ".join("?" for _ in cols)
    col_list = ", ".join(cols)
    con.executemany(
        f"INSERT INTO cases ({col_list}) VALUES ({placeholders})",
        [[r.get(c) for c in cols] for r in rows],
    )
    con.execute(CREATE_FTS)
    con.execute(POPULATE_FTS)
    con.commit()
    con.close()


if __name__ == "__main__":
    print(f"Loading {EXCEL.name} …")
    rows = load()
    print(f"  {len(rows)} cases loaded")
    print(f"Building {DB.name} …")
    build_db(rows)
    print("Done.")
